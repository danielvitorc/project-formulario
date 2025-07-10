from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl
from PIL import Image
import io
from ..models import Chamado

def preencher_excel_per(per):
    wb = load_workbook("formulario/static/per_modelo.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=30):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    ws["C5"] = per.nome_colaborador or ""
    ws["K5"] = per.matricula or ""
    ws["C6"] = per.funcao or ""
    ws["H6"] = per.depto or ""
    ws["C7"] = per.gestor_imediato or ""

    ws["B10"] = "[X]" if per.tipo_exposicao == "Exposição Intermitente" else "[  ]"

    ws["B13"] = "[X]" if per.natureza_risco == "Operações Perigosas Com Explosivos" else "[  ]"
    ws["B14"] = "[X]" if per.natureza_risco == "Operações Perigosas Com Inflamáveis" else "[  ]"
    ws["B15"] = "[X]" if per.natureza_risco == "Operações Perigosas Com Exposição A Roubos Ou Outras Espécies De Violência Física Nas Atividades Profissionais De Segurança Pessoal Ou Patrimonial" else "[  ]"
    ws["B16"] = "[X]" if per.natureza_risco == "Operações Perigosas Com Energia Elétrica" else "[  ]"
    ws["B17"] = "[X]" if per.natureza_risco == "Atividades Perigosas Em Motocicleta" else "[  ]"
    ws["B18"] = "[X]" if per.natureza_risco == "Atividades E Operações Perigosas Com Radiações Ionizantes Ou Substâncias Radiotivas" else "[  ]"

    ws["F9"] = per.descricao_atividades or ""
    ws["B20"] = "-"
    ws["B19"].alignment = Alignment(wrap_text=True)
    ws["F20"] = per.locais_atuaçao or ""
    ws["J20"] = per.frequencia or ""
    ws["B22"] = per.data_autorizacao_gestor.strftime("%d/%m/%Y") if per.data_autorizacao_gestor else ""
    ws["D22"] = per.responsavel or ""

    ws["H27"] = "[X]" if per.aso == "Apto" else "[  ]"
    ws["J27"] = "[X]" if per.aso == "Não apto" else "[  ]"
    ws["K27"] = "[X]" if per.aso == "Não aplicável" else "[  ]"

    ws["B27"] = f"Observações: {per.aso_descricao}" if per.aso_descricao else "Observações:"

    ws["H30"] = "[X]" if per.epi_epc == "Apto" else "[  ]"
    ws["J30"] = "[X]" if per.epi_epc == "Não apto" else "[  ]"
    ws["K30"] = "[X]" if per.epi_epc == "Não aplicável" else "[  ]"

    ws["B30"] = f"Observações: {per.epi_epc_descricao}" if per.epi_epc_descricao else "Observações:"

    ws["H33"] = "[X]" if per.curso_nr10 == "Apto" else "[  ]"
    ws["J33"] = "[X]" if per.curso_nr10 == "Não apto" else "[  ]"
    ws["K33"] = "[X]" if per.curso_nr10 == "Não aplicável" else "[  ]"

    ws["H34"] = "[X]" if per.curso_sep == "Apto" else "[  ]"
    ws["J34"] = "[X]" if per.curso_sep == "Não apto" else "[  ]"
    ws["K34"] = "[X]" if per.curso_sep == "Não aplicável" else "[  ]"

    ws["H35"] = "[X]" if per.curso_nr35 == "Apto" else "[  ]"
    ws["J35"] = "[X]" if per.curso_nr35 == "Não apto" else "[  ]"
    ws["K35"] = "[X]" if per.curso_nr35 == "Não aplicável" else "[  ]"

    ws["B36"] = f"Observações: {per.cursos_observacoes}" if per.cursos_observacoes else "Observações:"
   
    ws["B40"] = per.data_autorizacao_sesmt.strftime("%d/%m/%Y") if per.data_autorizacao_sesmt else ""
    ws["D40"] = per.nome_sesmt or ""

    ws["B45"] = "[X]" if per.procedimento_rh_dp == "Recebido sinalização automática da autorização" else "[  ]"
    ws["B46"] = "[X]" if per.procedimento_rh_dp == "Validação do relatório de comprovação via GPM" else "[  ]"
    ws["B47"] = "[X]" if per.procedimento_rh_dp == "Registro do adicional de periculosidade" else "[  ]"

    ws["B51"] = per.data_autorizacao_rh_dp.strftime("%d/%m/%Y") if per.data_autorizacao_rh_dp else ""
    ws["D51"] = per.nome_rh_dp or ""

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, per.imagem_assinatura_gestor, "J22")
    adicionar_imagem_excel(ws, per.imagem_assinatura_sesmt, "J40")
    adicionar_imagem_excel(ws, per.imagem_assinatura_rh_dp, "J51")

    # Salva em memória
    buffer = io.BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_per_excel(request, registro_id):
    registro = get_object_or_404(Chamado, id=registro_id)

    excel_bytes = preencher_excel_per(registro)
    nome_arquivo = f"PER_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


@login_required
def exportar_chamados_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chamados"

    # Cabeçalhos
    colunas = [
        "Nome", "Matrícula", "Função", "Depto", "Gestor Imediato", "Tipo de Exposição",
        "Natureza do Risco", "Descrição das Atividades", "Locais de Atuação", "Frequência", "Responsável",
        "Data Aut. Gestor", "ASO", "ASO Desc.", "EPI/EPC", "EPI/EPC Desc.",
        "Curso NR10", "Curso SEP", "Curso NR35", "Cursos Obs.", "Data Aut. SESMT", "Nome SESMT",
        "Procedimento RH/DP", "Data Aut. RH/DP", "Nome RH/DP"
    ]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escrever cabeçalhos com estilo
    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Dados
    chamados = Chamado.objects.all()

    for row_num, chamado in enumerate(chamados, 2):
        dados = [
            chamado.nome_colaborador,
            chamado.matricula,
            chamado.funcao,
            chamado.depto,
            chamado.gestor_imediato,
            chamado.tipo_exposicao,
            chamado.natureza_risco,
            chamado.descricao_atividades,
            chamado.locais_atuaçao,
            chamado.frequencia,
            chamado.responsavel,
            chamado.data_autorizacao_gestor,
            chamado.aso,
            chamado.aso_descricao,
            chamado.epi_epc,
            chamado.epi_epc_descricao,
            chamado.curso_nr10,
            chamado.curso_sep,
            chamado.curso_nr35,
            chamado.cursos_observacoes,
            chamado.data_autorizacao_sesmt,
            chamado.nome_sesmt,
            chamado.procedimento_rh_dp,
            chamado.data_autorizacao_rh_dp,
            chamado.nome_rh_dp,
        ]

        for col_num, valor in enumerate(dados, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Resposta HTTP com o arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=chamados_exportados.xlsx'
    wb.save(response)
    return response